
from __future__ import annotations

import io
from datetime import datetime
from typing import Mapping

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from release_viewer import release_detail_rows
from dashboard import action_items, confidence_distribution, status_counts, vehicle_summary

COLORS = {
    "MATCH": "C6EFCE", "COMPLIANT": "C6EFCE",
    "UPDATE_AVAILABLE": "FFF2CC", "PART_MATCH": "FFEB9C", "PARTIAL_MATCH": "FFEB9C",
    "REVIEW": "FCE4D6", "MISSING": "FCE4D6",
    "MISMATCH": "FFC7CE", "WRONG_RELEASE": "F4B084",
    "NO_REFERENCE": "D9E1F2", "NOT_APPLICABLE": "E7E6E6",
    "HIGH": "FFC7CE", "MEDIUM": "FFF2CC", "LOW": "D9EAF7",
}
HEADER = "1F4E78"
SUBHEADER = "D9EAF7"
TEXT_DARK = "1F2937"
BORDER_COLOR = "D9E1F2"


def _safe(value):
    if isinstance(value, tuple):
        return ".".join(map(str, value))
    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass
    return value.item() if hasattr(value, "item") else value


def _style_header(ws, row=1):
    for cell in ws[row]:
        cell.fill = PatternFill("solid", fgColor=HEADER)
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _add_sheet(wb: Workbook, name: str, df: pd.DataFrame, status_col: str | None = None):
    ws = wb.create_sheet(name[:31])
    ws.append(list(df.columns))
    for row in df.itertuples(index=False, name=None):
        ws.append([_safe(v) for v in row])
    _style_header(ws)
    ws.freeze_panes = "A2"
    if ws.max_row > 1:
        ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    thin = Side(style="thin", color=BORDER_COLOR)
    status_idx = list(df.columns).index(status_col) + 1 if status_col and status_col in df.columns else None
    for r in range(2, ws.max_row + 1):
        status = str(ws.cell(r, status_idx).value or "") if status_idx else ""
        fill = PatternFill("solid", fgColor=COLORS.get(status, "FFFFFF"))
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(r, c)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(bottom=thin)
            if status_idx:
                cell.fill = fill
    for c in range(1, ws.max_column + 1):
        width = max([len(str(ws.cell(r, c).value or "")) for r in range(1, min(ws.max_row, 300) + 1)] + [10]) + 2
        ws.column_dimensions[get_column_letter(c)].width = min(max(width, 11), 48)
    ws.sheet_view.showGridLines = False
    return ws


def _executive_summary_sheet(
    wb: Workbook,
    overview: pd.DataFrame,
    release_sheet: str,
    metadata: Mapping[str, str],
):
    ws = wb.create_sheet("Executive Summary")
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:H2")
    ws["A1"] = metadata.get("report_title") or "Grade-X Software Compliance Report"
    ws["A1"].font = Font(size=20, bold=True, color=HEADER)
    ws["A1"].alignment = Alignment(vertical="center")

    ws["A4"] = "Target release"
    ws["B4"] = release_sheet
    ws["D4"] = "Prepared by"
    ws["E4"] = metadata.get("prepared_by", "")
    ws["A5"] = "Department / Project"
    ws["B5"] = metadata.get("project", "")
    ws["D5"] = "Report date"
    ws["E5"] = metadata.get("report_date", datetime.now().strftime("%Y-%m-%d %H:%M"))
    ws["A6"] = "Reference file"
    ws["B6"] = metadata.get("reference_file", "")
    ws["D6"] = "Vehicles / sessions"
    ws["E6"] = int(overview["Source File"].nunique()) if "Source File" in overview else 0

    for cell in ("A4", "D4", "A5", "D5", "A6", "D6"):
        ws[cell].font = Font(bold=True, color=TEXT_DARK)
        ws[cell].fill = PatternFill("solid", fgColor=SUBHEADER)

    statuses = overview.get("Status", pd.Series(dtype=str)).astype(str)
    total = len(overview)
    compliant = int((statuses == "COMPLIANT").sum())
    critical = int(statuses.isin(["MISMATCH", "WRONG_RELEASE"]).sum())
    updates = int((statuses == "UPDATE_AVAILABLE").sum())
    persistent_dtcs = int(pd.to_numeric(
        overview.get("Persistent DTC Count", pd.Series(dtype=float)), errors="coerce"
    ).fillna(0).sum())
    avg_conf = pd.to_numeric(overview.get("Confidence %", pd.Series(dtype=float)), errors="coerce").mean()
    compliance_rate = compliant / total * 100 if total else 0

    metrics = [
        ("ECUs checked", total, "D9EAF7"),
        ("Compliance rate", f"{compliance_rate:.1f}%", "C6EFCE"),
        ("Critical findings", critical, "FFC7CE" if critical else "C6EFCE"),
        ("Updates available", updates, "FFF2CC"),
        ("Persistent DTCs", persistent_dtcs, "FCE4D6" if persistent_dtcs else "C6EFCE"),
    ]
    start_cols = [1, 3, 5, 7, 9]
    for (label, value, color), col in zip(metrics, start_cols):
        ws.merge_cells(start_row=8, start_column=col, end_row=8, end_column=col+1)
        ws.merge_cells(start_row=9, start_column=col, end_row=10, end_column=col+1)
        label_cell = ws.cell(8, col)
        value_cell = ws.cell(9, col)
        label_cell.value = label
        value_cell.value = value
        label_cell.font = Font(bold=True, color=TEXT_DARK)
        label_cell.alignment = Alignment(horizontal="center")
        value_cell.font = Font(size=18, bold=True, color=HEADER)
        value_cell.alignment = Alignment(horizontal="center", vertical="center")
        for row in range(8, 11):
            for c in range(col, col+2):
                ws.cell(row, c).fill = PatternFill("solid", fgColor=color)
                ws.cell(row, c).border = Border(
                    left=Side(style="thin", color=BORDER_COLOR),
                    right=Side(style="thin", color=BORDER_COLOR),
                    top=Side(style="thin", color=BORDER_COLOR),
                    bottom=Side(style="thin", color=BORDER_COLOR),
                )

    ws["A13"] = "Status"
    ws["B13"] = "Count"
    ws["C13"] = "Share"
    status_order = [
        "COMPLIANT", "UPDATE_AVAILABLE", "PARTIAL_MATCH", "REVIEW",
        "MISMATCH", "WRONG_RELEASE", "NO_REFERENCE"
    ]
    for idx, status in enumerate(status_order, start=14):
        count = int((statuses == status).sum())
        ws.cell(idx, 1, status)
        ws.cell(idx, 2, count)
        ws.cell(idx, 3, count / total if total else 0)
        ws.cell(idx, 3).number_format = "0.0%"
        for c in range(1, 4):
            ws.cell(idx, c).fill = PatternFill("solid", fgColor=COLORS.get(status, "FFFFFF"))
    for c in range(1, 4):
        ws.cell(13, c).fill = PatternFill("solid", fgColor=HEADER)
        ws.cell(13, c).font = Font(color="FFFFFF", bold=True)

    pie = PieChart()
    labels = Reference(ws, min_col=1, min_row=14, max_row=13 + len(status_order))
    values = Reference(ws, min_col=2, min_row=13, max_row=13 + len(status_order))
    pie.add_data(values, titles_from_data=True)
    pie.set_categories(labels)
    pie.title = "Compliance status distribution"
    pie.height = 8.5
    pie.width = 12
    ws.add_chart(pie, "E13")

    if metadata.get("executive_comment"):
        ws.merge_cells("A24:J24")
        ws["A24"] = "Executive comment"
        ws["A24"].fill = PatternFill("solid", fgColor=HEADER)
        ws["A24"].font = Font(color="FFFFFF", bold=True)
        ws.merge_cells("A25:J30")
        ws["A25"] = metadata["executive_comment"]
        ws["A25"].alignment = Alignment(wrap_text=True, vertical="top")
        ws["A25"].fill = PatternFill("solid", fgColor="F8FAFC")

    for col in range(1, 11):
        ws.column_dimensions[get_column_letter(col)].width = 15
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["E"].width = 25
    ws.freeze_panes = "A4"
    return ws


def _vehicle_report_sheets(wb: Workbook, overview: pd.DataFrame, details: pd.DataFrame):
    if "Source File" not in overview.columns:
        return
    for index, (source, group) in enumerate(overview.groupby("Source File", dropna=False), start=1):
        name = f"Vehicle {index}"[:31]
        ws = wb.create_sheet(name)
        ws.sheet_view.showGridLines = False
        ws.merge_cells("A1:J2")
        ws["A1"] = f"Vehicle / Session: {source}"
        ws["A1"].font = Font(size=16, bold=True, color=HEADER)
        vin_series = group.get("VIN", pd.Series(dtype=str)).dropna()
        ws["A4"] = "VIN"
        ws["B4"] = vin_series.iloc[0] if not vin_series.empty else ""
        ws["D4"] = "ECUs"
        ws["E4"] = len(group)
        ws["G4"] = "Compliance rate"
        ws["H4"] = f"{(group['Status'].astype(str) == 'COMPLIANT').mean() * 100:.1f}%"

        display_cols = [
            "ECU", "Status", "Matched Variant", "Installed Release", "Target Release",
            "Confidence %", "Matches", "Partial Matches", "Mismatches", "Missing", "Decision Reason"
        ]
        display_cols = [c for c in display_cols if c in group.columns]
        start_row = 6
        for c, header in enumerate(display_cols, start=1):
            ws.cell(start_row, c, header)
        _style_header(ws, start_row)
        for r_idx, (_, row) in enumerate(group[display_cols].iterrows(), start=start_row+1):
            status = str(row.get("Status", ""))
            fill = PatternFill("solid", fgColor=COLORS.get(status, "FFFFFF"))
            for c_idx, col in enumerate(display_cols, start=1):
                cell = ws.cell(r_idx, c_idx, _safe(row[col]))
                cell.fill = fill
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        source_details = details[details["Source File"] == source] if "Source File" in details.columns else pd.DataFrame()
        deviations = source_details[source_details["Status"].astype(str).isin(["MISMATCH", "MISSING", "PART_MATCH"])] if not source_details.empty else pd.DataFrame()
        next_row = ws.max_row + 3
        if not deviations.empty:
            ws.cell(next_row, 1, "Identifier Deviations")
            ws.cell(next_row, 1).font = Font(size=13, bold=True, color=HEADER)
            next_row += 1
            dev_cols = [c for c in ["ECU", "Field", "Actual", "Expected", "Status", "Reason"] if c in deviations.columns]
            for c_idx, col in enumerate(dev_cols, start=1):
                ws.cell(next_row, c_idx, col)
            _style_header(ws, next_row)
            for r_idx, (_, row) in enumerate(deviations[dev_cols].iterrows(), start=next_row+1):
                fill = PatternFill("solid", fgColor=COLORS.get(str(row.get("Status", "")), "FFFFFF"))
                for c_idx, col in enumerate(dev_cols, start=1):
                    ws.cell(r_idx, c_idx, _safe(row[col])).fill = fill

        for c in range(1, ws.max_column + 1):
            width = max([len(str(ws.cell(r, c).value or "")) for r in range(1, min(ws.max_row, 250)+1)] + [10]) + 2
            ws.column_dimensions[get_column_letter(c)].width = min(max(width, 11), 46)
        ws.freeze_panes = "A7"


def build_report(
    overview: pd.DataFrame,
    details: pd.DataFrame,
    summary: pd.DataFrame,
    raw: pd.DataFrame,
    candidates: pd.DataFrame,
    release_sheet: str,
    metadata: Mapping[str, str] | None = None,
    dtc_summary: pd.DataFrame | None = None,
    dtc_events: pd.DataFrame | None = None,
    vehicle_history: pd.DataFrame | None = None,
    change_log: pd.DataFrame | None = None,
    fleet_overview: pd.DataFrame | None = None,
    risk_breakdown: pd.DataFrame | None = None,
    vehicle_health: pd.DataFrame | None = None,
    risk_rules: Mapping[str, object] | None = None,
    decision_actions: pd.DataFrame | None = None,
    vehicle_decisions: pd.DataFrame | None = None,
    decision_rules: Mapping[str, object] | None = None,
    warranty_summary: pd.DataFrame | None = None,
    ecu_warranty: pd.DataFrame | None = None,
    engineering_summaries: pd.DataFrame | None = None,
    warranty_rules: Mapping[str, object] | None = None,
    dependency_nodes: pd.DataFrame | None = None,
    dependency_edges: pd.DataFrame | None = None,
    vehicle_root_causes: pd.DataFrame | None = None,
    root_cause_ranking: pd.DataFrame | None = None,
    root_cause_paths: pd.DataFrame | None = None,
    dependency_rules: Mapping[str, object] | None = None,
    release_consistency: pd.DataFrame | None = None,
    consistency_fields: pd.DataFrame | None = None,
    ecu_consistency: pd.DataFrame | None = None,
    consistency_rules: Mapping[str, object] | None = None,
    fleet_kpis: pd.DataFrame | None = None,
    fleet_problematic_ecus: pd.DataFrame | None = None,
    fleet_dtc_ranking: pd.DataFrame | None = None,
    fleet_release_patterns: pd.DataFrame | None = None,
    fleet_root_ecus: pd.DataFrame | None = None,
    fleet_warranty_patterns: pd.DataFrame | None = None,
    fleet_alerts: pd.DataFrame | None = None,
    fleet_rules: Mapping[str, object] | None = None,
    engineering_findings: pd.DataFrame | None = None,
    assistant_summary: pd.DataFrame | None = None,
    assistant_action_plan: pd.DataFrame | None = None,
    assistant_rules: Mapping[str, object] | None = None,
    workspace_timeseries: pd.DataFrame | None = None,
    timeseries_transitions: pd.DataFrame | None = None,
    vehicle_trend_summary: pd.DataFrame | None = None,
    timeline_events: pd.DataFrame | None = None,
    timeseries_rules: Mapping[str, object] | None = None,
    network_nodes: pd.DataFrame | None = None,
    network_edges: pd.DataFrame | None = None,
    dependency_matrix: pd.DataFrame | None = None,
    network_violations: pd.DataFrame | None = None,
    network_health: pd.DataFrame | None = None,
    network_criticality: pd.DataFrame | None = None,
    network_rules: Mapping[str, object] | None = None,
    network_statistics: pd.DataFrame | None = None,
    network_components: pd.DataFrame | None = None,
    network_cycles: pd.DataFrame | None = None,
    network_heatmap: pd.DataFrame | None = None,
    network_explorer: pd.DataFrame | None = None,
    update_impact_summary: pd.DataFrame | None = None,
    update_impact_results: pd.DataFrame | None = None,
    companion_updates: pd.DataFrame | None = None,
    update_sequence: pd.DataFrame | None = None,
    update_impact_rules: Mapping[str, object] | None = None,
    release_path_summary: pd.DataFrame | None = None,
    release_scenario_comparison: pd.DataFrame | None = None,
    release_scenario_impacts: pd.DataFrame | None = None,
    release_scenario_companions: pd.DataFrame | None = None,
    release_scenario_sequences: pd.DataFrame | None = None,
    network_timeline: pd.DataFrame | None = None,
    release_path_rules: Mapping[str, object] | None = None,
    release_recommendation_summary: pd.DataFrame | None = None,
    release_recommendation_ranking: pd.DataFrame | None = None,
    historical_release_performance: pd.DataFrame | None = None,
    release_update_plan: pd.DataFrame | None = None,
    release_recommendation_rules: Mapping[str, object] | None = None,
    configuration_diff_summary: pd.DataFrame | None = None,
    configuration_diff_ecus: pd.DataFrame | None = None,
    configuration_diff_fields: pd.DataFrame | None = None,
    multi_session_catalog: pd.DataFrame | None = None,
    multi_session_transition_summary: pd.DataFrame | None = None,
    multi_session_transition_ecus: pd.DataFrame | None = None,
    multi_session_transition_fields: pd.DataFrame | None = None,
    multi_session_ecu_timeline: pd.DataFrame | None = None,
    multi_session_change_history: pd.DataFrame | None = None,
    multi_session_vehicle_trend: pd.DataFrame | None = None,
    advanced_search_summary: pd.DataFrame | None = None,
    advanced_search_results: pd.DataFrame | None = None,
    compliance_dashboard_summary: pd.DataFrame | None = None,
    compliance_dashboard_status: pd.DataFrame | None = None,
    compliance_dashboard_risk: pd.DataFrame | None = None,
    compliance_dashboard_failing: pd.DataFrame | None = None,
    compliance_dashboard_release: pd.DataFrame | None = None,
    compliance_dashboard_trend: pd.DataFrame | None = None,
    release_coverage_summary: pd.DataFrame | None = None,
    release_coverage_ecus: pd.DataFrame | None = None,
    release_coverage_installed: pd.DataFrame | None = None,
    release_coverage_target: pd.DataFrame | None = None,
    release_coverage_by_target: pd.DataFrame | None = None,
    release_coverage_legacy: pd.DataFrame | None = None,
    release_coverage_unknown: pd.DataFrame | None = None,
    release_reference_coverage: pd.DataFrame | None = None,
    release_session_trend: pd.DataFrame | None = None,
    oem_audit_summary: pd.DataFrame | None = None,
    oem_audit_checklist: pd.DataFrame | None = None,
    oem_audit_findings: pd.DataFrame | None = None,
    oem_audit_area_summary: pd.DataFrame | None = None,
    oem_audit_rules: Mapping[str, object] | None = None,
    session_merge_summary: pd.DataFrame | None = None,
    session_merge_snapshot: pd.DataFrame | None = None,
    session_merge_provenance: pd.DataFrame | None = None,
    session_merge_conflicts: pd.DataFrame | None = None,
    session_merge_presence: pd.DataFrame | None = None,
    session_merge_catalog: pd.DataFrame | None = None,
    session_merge_history: pd.DataFrame | None = None,
) -> bytes:
    metadata = dict(metadata or {})
    wb = Workbook()
    wb.remove(wb.active)

    _executive_summary_sheet(wb, overview, release_sheet, metadata)
    _add_sheet(wb, "Dashboard Status", status_counts(overview), "Status")
    _add_sheet(wb, "Vehicle Summary", vehicle_summary(overview))
    _add_sheet(wb, "Priority Actions", action_items(overview, details), "Status")
    _add_sheet(wb, "Confidence Distribution", confidence_distribution(overview))
    _add_sheet(wb, "Compliance Overview", overview, "Status")
    _add_sheet(wb, "Release Details", release_detail_rows(summary, overview, details), "Status")
    _add_sheet(wb, "Field Details", details, "Status")
    _add_sheet(wb, "ECU Identifiers", summary)
    _add_sheet(wb, "Candidate Variants", candidates)
    _add_sheet(wb, "Raw Identifiers", raw)
    if dtc_summary is not None and not dtc_summary.empty:
        _add_sheet(wb, "DTC Summary", dtc_summary, "Severity")
    if dtc_events is not None and not dtc_events.empty:
        _add_sheet(wb, "DTC Events", dtc_events)
    if fleet_overview is not None and not fleet_overview.empty:
        _add_sheet(wb, "Fleet Summary", fleet_overview)
    if change_log is not None and not change_log.empty:
        _add_sheet(wb, "Change Log", change_log)
    if vehicle_history is not None and not vehicle_history.empty:
        _add_sheet(wb, "Vehicle History", vehicle_history)
    if vehicle_health is not None and not vehicle_health.empty:
        _add_sheet(wb, "Vehicle Health", vehicle_health)
    if risk_breakdown is not None and not risk_breakdown.empty:
        _add_sheet(wb, "Risk Breakdown", risk_breakdown, "Risk Level")
    if risk_rules:
        rules_rows = []
        for section, values in risk_rules.items():
            if isinstance(values, dict):
                for key, value in values.items():
                    rules_rows.append({"Section": section, "Rule": key, "Value": value})
            else:
                rules_rows.append({"Section": "General", "Rule": section, "Value": values})
        _add_sheet(wb, "Risk Rules", pd.DataFrame(rules_rows))
    if vehicle_decisions is not None and not vehicle_decisions.empty:
        _add_sheet(wb, "Vehicle Decisions", vehicle_decisions, "Overall Urgency")
    if decision_actions is not None and not decision_actions.empty:
        _add_sheet(wb, "Decision Actions", decision_actions, "Decision Urgency")
    if decision_rules:
        decision_rule_rows = []
        for section, values in decision_rules.items():
            if isinstance(values, dict):
                for key, value in values.items():
                    decision_rule_rows.append({
                        "Section": section,
                        "Rule": key,
                        "Value": str(value),
                    })
            else:
                decision_rule_rows.append({
                    "Section": "General",
                    "Rule": section,
                    "Value": str(values),
                })
        _add_sheet(wb, "Decision Rules", pd.DataFrame(decision_rule_rows))
    if warranty_summary is not None and not warranty_summary.empty:
        _add_sheet(wb, "Warranty Triage", warranty_summary)
    if ecu_warranty is not None and not ecu_warranty.empty:
        _add_sheet(wb, "ECU Warranty Triage", ecu_warranty)
    if engineering_summaries is not None and not engineering_summaries.empty:
        _add_sheet(wb, "Engineering Summaries", engineering_summaries)
    if warranty_rules:
        warranty_rule_rows = []
        for section, values in warranty_rules.items():
            if isinstance(values, dict):
                for key, value in values.items():
                    warranty_rule_rows.append({
                        "Section": section,
                        "Rule": key,
                        "Value": str(value),
                    })
            else:
                warranty_rule_rows.append({
                    "Section": "General",
                    "Rule": section,
                    "Value": str(values),
                })
        _add_sheet(wb, "Warranty Rules", pd.DataFrame(warranty_rule_rows))
    if dependency_nodes is not None and not dependency_nodes.empty:
        _add_sheet(wb, "Dependency Nodes", dependency_nodes)
    if dependency_edges is not None and not dependency_edges.empty:
        _add_sheet(wb, "Dependency Edges", dependency_edges)
    if vehicle_root_causes is not None and not vehicle_root_causes.empty:
        _add_sheet(wb, "Vehicle Root Causes", vehicle_root_causes)
    if root_cause_ranking is not None and not root_cause_ranking.empty:
        _add_sheet(wb, "Root Cause Ranking", root_cause_ranking)
    if root_cause_paths is not None and not root_cause_paths.empty:
        _add_sheet(wb, "Root Cause Paths", root_cause_paths)
    if dependency_rules:
        dependency_rule_rows = []
        for section, values in dependency_rules.items():
            if isinstance(values, dict):
                for key, value in values.items():
                    dependency_rule_rows.append({
                        "Section": section,
                        "Rule": key,
                        "Value": str(value),
                    })
            elif isinstance(values, list):
                for index, value in enumerate(values, start=1):
                    dependency_rule_rows.append({
                        "Section": section,
                        "Rule": index,
                        "Value": str(value),
                    })
            else:
                dependency_rule_rows.append({
                    "Section": "General",
                    "Rule": section,
                    "Value": str(values),
                })
        _add_sheet(wb, "Dependency Rules", pd.DataFrame(dependency_rule_rows))
    if release_consistency is not None and not release_consistency.empty:
        _add_sheet(wb, "Release Consistency", release_consistency)
    if consistency_fields is not None and not consistency_fields.empty:
        _add_sheet(wb, "Consistency Fields", consistency_fields)
    if ecu_consistency is not None and not ecu_consistency.empty:
        _add_sheet(wb, "ECU Consistency", ecu_consistency)
    if consistency_rules:
        consistency_rule_rows = []
        for section, values in consistency_rules.items():
            if isinstance(values, dict):
                for key, value in values.items():
                    consistency_rule_rows.append({
                        "Section": section,
                        "Rule": key,
                        "Value": str(value),
                    })
            elif isinstance(values, list):
                for index, value in enumerate(values, start=1):
                    consistency_rule_rows.append({
                        "Section": section,
                        "Rule": index,
                        "Value": str(value),
                    })
            else:
                consistency_rule_rows.append({
                    "Section": "General",
                    "Rule": section,
                    "Value": str(values),
                })
        _add_sheet(wb, "Consistency Rules", pd.DataFrame(consistency_rule_rows))
    if fleet_kpis is not None and not fleet_kpis.empty:
        _add_sheet(wb, "Fleet KPIs", fleet_kpis)
    if fleet_problematic_ecus is not None and not fleet_problematic_ecus.empty:
        _add_sheet(wb, "Fleet ECU Ranking", fleet_problematic_ecus)
    if fleet_dtc_ranking is not None and not fleet_dtc_ranking.empty:
        _add_sheet(wb, "Fleet DTC Ranking", fleet_dtc_ranking)
    if fleet_release_patterns is not None and not fleet_release_patterns.empty:
        _add_sheet(wb, "Fleet Release Patterns", fleet_release_patterns)
    if fleet_root_ecus is not None and not fleet_root_ecus.empty:
        _add_sheet(wb, "Fleet Root ECUs", fleet_root_ecus)
    if fleet_warranty_patterns is not None and not fleet_warranty_patterns.empty:
        _add_sheet(wb, "Fleet Warranty Patterns", fleet_warranty_patterns)
    if fleet_alerts is not None and not fleet_alerts.empty:
        _add_sheet(wb, "Fleet Alerts", fleet_alerts)
    if fleet_rules:
        fleet_rule_rows = []
        for section, values in fleet_rules.items():
            if isinstance(values, dict):
                for key, value in values.items():
                    fleet_rule_rows.append({
                        "Section": section,
                        "Rule": key,
                        "Value": str(value),
                    })
            else:
                fleet_rule_rows.append({
                    "Section": "General",
                    "Rule": section,
                    "Value": str(values),
                })
        _add_sheet(wb, "Fleet Rules", pd.DataFrame(fleet_rule_rows))
    if engineering_findings is not None and not engineering_findings.empty:
        _add_sheet(wb, "Engineering Findings", engineering_findings, "Severity")
    if assistant_summary is not None and not assistant_summary.empty:
        _add_sheet(wb, "Executive Summaries", assistant_summary)
    if assistant_action_plan is not None and not assistant_action_plan.empty:
        _add_sheet(wb, "Assistant Action Plan", assistant_action_plan, "Priority")
    if assistant_rules:
        assistant_rule_rows = []
        for section, values in assistant_rules.items():
            if isinstance(values, dict):
                for key, value in values.items():
                    assistant_rule_rows.append({
                        "Section": section,
                        "Rule": key,
                        "Value": str(value),
                    })
            elif isinstance(values, list):
                for index, value in enumerate(values, start=1):
                    assistant_rule_rows.append({
                        "Section": section,
                        "Rule": index,
                        "Value": str(value),
                    })
            else:
                assistant_rule_rows.append({
                    "Section": "General",
                    "Rule": section,
                    "Value": str(values),
                })
        _add_sheet(wb, "Assistant Rules", pd.DataFrame(assistant_rule_rows))
    if workspace_timeseries is not None and not workspace_timeseries.empty:
        _add_sheet(wb, "Workspace Time Series", workspace_timeseries)
    if timeseries_transitions is not None and not timeseries_transitions.empty:
        _add_sheet(wb, "Time Series Transitions", timeseries_transitions)
    if vehicle_trend_summary is not None and not vehicle_trend_summary.empty:
        _add_sheet(wb, "Vehicle Trend Summary", vehicle_trend_summary)
    if timeline_events is not None and not timeline_events.empty:
        _add_sheet(wb, "Timeline Events", timeline_events)
    if timeseries_rules:
        timeseries_rule_rows = []
        for section, values in timeseries_rules.items():
            if isinstance(values, dict):
                for key, value in values.items():
                    timeseries_rule_rows.append({
                        "Section": section,
                        "Rule": key,
                        "Value": str(value),
                    })
            else:
                timeseries_rule_rows.append({
                    "Section": "General",
                    "Rule": section,
                    "Value": str(values),
                })
        _add_sheet(wb, "Time Series Rules", pd.DataFrame(timeseries_rule_rows))
    if network_nodes is not None and not network_nodes.empty:
        _add_sheet(wb, "ECU Network Nodes", network_nodes)
    if network_edges is not None and not network_edges.empty:
        _add_sheet(wb, "ECU Network Edges", network_edges)
    if dependency_matrix is not None and not dependency_matrix.empty:
        _add_sheet(wb, "Dependency Matrix", dependency_matrix)
    if network_violations is not None and not network_violations.empty:
        _add_sheet(wb, "Network Violations", network_violations, "Violation Severity")
    if network_health is not None and not network_health.empty:
        _add_sheet(wb, "Network Health", network_health, "Network Health Level")
    if network_criticality is not None and not network_criticality.empty:
        _add_sheet(wb, "ECU Criticality", network_criticality, "Criticality")
    if network_statistics is not None and not network_statistics.empty:
        _add_sheet(wb, "Network Statistics", network_statistics, "Graph Integrity Level")
    if network_components is not None and not network_components.empty:
        _add_sheet(wb, "Connected Components", network_components)
    if network_cycles is not None and not network_cycles.empty:
        _add_sheet(wb, "Circular Dependencies", network_cycles, "Severity")
    if network_heatmap is not None and not network_heatmap.empty:
        _add_sheet(wb, "Network Heatmap", network_heatmap, "Criticality")
    if network_explorer is not None and not network_explorer.empty:
        _add_sheet(wb, "Network Explorer", network_explorer, "Criticality")
    if update_impact_summary is not None and not update_impact_summary.empty:
        _add_sheet(wb, "Update Impact Summary", update_impact_summary, "Overall Compatibility Level")
    if update_impact_results is not None and not update_impact_results.empty:
        _add_sheet(wb, "Update Impact Results", update_impact_results, "Compatibility Level")
    if companion_updates is not None and not companion_updates.empty:
        _add_sheet(wb, "Companion ECU Updates", companion_updates, "Compatibility Level")
    if update_sequence is not None and not update_sequence.empty:
        _add_sheet(wb, "Update Sequence", update_sequence, "Compatibility Level")
    if release_path_summary is not None and not release_path_summary.empty:
        _add_sheet(wb, "Release Path Summary", release_path_summary, "Path Type")
    if release_scenario_comparison is not None and not release_scenario_comparison.empty:
        _add_sheet(wb, "Release Scenarios", release_scenario_comparison, "Path Type")
    if release_scenario_impacts is not None and not release_scenario_impacts.empty:
        _add_sheet(wb, "Scenario Impact Details", release_scenario_impacts, "Compatibility Level")
    if release_scenario_companions is not None and not release_scenario_companions.empty:
        _add_sheet(wb, "Scenario Companion Updates", release_scenario_companions, "Compatibility Level")
    if release_scenario_sequences is not None and not release_scenario_sequences.empty:
        _add_sheet(wb, "Scenario Update Sequences", release_scenario_sequences, "Compatibility Level")
    if network_timeline is not None and not network_timeline.empty:
        _add_sheet(wb, "Network Timeline", network_timeline, "Root ECU Criticality")
    if release_recommendation_summary is not None and not release_recommendation_summary.empty:
        _add_sheet(wb, "Release Recommendation", release_recommendation_summary, "Recommendation Decision")
    if release_recommendation_ranking is not None and not release_recommendation_ranking.empty:
        _add_sheet(wb, "Recommendation Ranking", release_recommendation_ranking, "Recommendation Decision")
    if historical_release_performance is not None and not historical_release_performance.empty:
        _add_sheet(wb, "Historical Release Results", historical_release_performance)
    if release_update_plan is not None and not release_update_plan.empty:
        _add_sheet(wb, "Recommended Update Plan", release_update_plan, "Phase")
    if session_merge_summary is not None and not session_merge_summary.empty:
        _add_sheet(wb, "Session Merge Summary", session_merge_summary, "Merge Quality Level")
    if session_merge_snapshot is not None and not session_merge_snapshot.empty:
        _add_sheet(wb, "Unified Vehicle Snapshot", session_merge_snapshot, "Merge Status")
    if session_merge_provenance is not None and not session_merge_provenance.empty:
        _add_sheet(wb, "Merge Field Provenance", session_merge_provenance)
    if session_merge_conflicts is not None and not session_merge_conflicts.empty:
        _add_sheet(wb, "Session Merge Conflicts", session_merge_conflicts)
    if session_merge_presence is not None and not session_merge_presence.empty:
        _add_sheet(wb, "Merged ECU Presence", session_merge_presence)
    if session_merge_catalog is not None and not session_merge_catalog.empty:
        _add_sheet(wb, "Merged Session Catalog", session_merge_catalog)
    if session_merge_history is not None and not session_merge_history.empty:
        _add_sheet(wb, "Merged Change History", session_merge_history)
    if oem_audit_summary is not None and not oem_audit_summary.empty:
        _add_sheet(wb, "OEM Audit Summary", oem_audit_summary, "Audit Decision")
    if oem_audit_checklist is not None and not oem_audit_checklist.empty:
        _add_sheet(wb, "OEM Audit Checklist", oem_audit_checklist, "Status")
    if oem_audit_findings is not None and not oem_audit_findings.empty:
        _add_sheet(wb, "Open Audit Findings", oem_audit_findings, "Status")
    if oem_audit_area_summary is not None and not oem_audit_area_summary.empty:
        _add_sheet(wb, "Audit Area Summary", oem_audit_area_summary)
    if oem_audit_rules:
        audit_rule_rows = []
        for section, values in oem_audit_rules.items():
            if isinstance(values, dict):
                for key, value in values.items():
                    audit_rule_rows.append({
                        "Section": section, "Rule": key, "Value": str(value)
                    })
            elif isinstance(values, list):
                for index, value in enumerate(values, start=1):
                    audit_rule_rows.append({
                        "Section": section, "Rule": index, "Value": str(value)
                    })
            else:
                audit_rule_rows.append({
                    "Section": "General", "Rule": section, "Value": str(values)
                })
        _add_sheet(wb, "OEM Audit Rules", pd.DataFrame(audit_rule_rows))
    if release_coverage_summary is not None and not release_coverage_summary.empty:
        _add_sheet(wb, "Release Coverage Summary", release_coverage_summary, "Coverage Level")
    if release_coverage_ecus is not None and not release_coverage_ecus.empty:
        _add_sheet(wb, "ECU Release Coverage", release_coverage_ecus, "Coverage State")
    if release_coverage_installed is not None and not release_coverage_installed.empty:
        _add_sheet(wb, "Installed Release Distribution", release_coverage_installed)
    if release_coverage_target is not None and not release_coverage_target.empty:
        _add_sheet(wb, "Target Release Distribution", release_coverage_target)
    if release_coverage_by_target is not None and not release_coverage_by_target.empty:
        _add_sheet(wb, "Coverage by Target", release_coverage_by_target)
    if release_coverage_legacy is not None and not release_coverage_legacy.empty:
        _add_sheet(wb, "Legacy Release Candidates", release_coverage_legacy, "Coverage State")
    if release_coverage_unknown is not None and not release_coverage_unknown.empty:
        _add_sheet(wb, "Unknown Releases", release_coverage_unknown, "Coverage State")
    if release_reference_coverage is not None and not release_reference_coverage.empty:
        _add_sheet(wb, "Reference Coverage", release_reference_coverage, "Coverage Status")
    if release_session_trend is not None and not release_session_trend.empty:
        _add_sheet(wb, "Session Release Trend", release_session_trend)
    if compliance_dashboard_summary is not None and not compliance_dashboard_summary.empty:
        _add_sheet(wb, "Compliance Dashboard", compliance_dashboard_summary, "Overall Quality Level")
    if compliance_dashboard_status is not None and not compliance_dashboard_status.empty:
        _add_sheet(wb, "Compliance Distribution", compliance_dashboard_status)
    if compliance_dashboard_risk is not None and not compliance_dashboard_risk.empty:
        _add_sheet(wb, "Risk Distribution", compliance_dashboard_risk)
    if compliance_dashboard_failing is not None and not compliance_dashboard_failing.empty:
        _add_sheet(wb, "Top Failing ECUs", compliance_dashboard_failing, "Status")
    if compliance_dashboard_release is not None and not compliance_dashboard_release.empty:
        _add_sheet(wb, "Release Distribution", compliance_dashboard_release)
    if compliance_dashboard_trend is not None and not compliance_dashboard_trend.empty:
        _add_sheet(wb, "Quality Transition Trend", compliance_dashboard_trend, "Overall Assessment")
    if advanced_search_summary is not None and not advanced_search_summary.empty:
        _add_sheet(wb, "Advanced Search Summary", advanced_search_summary)
    if advanced_search_results is not None and not advanced_search_results.empty:
        _add_sheet(wb, "Advanced Search Results", advanced_search_results)
    if multi_session_catalog is not None and not multi_session_catalog.empty:
        _add_sheet(wb, "Multi Session Catalog", multi_session_catalog)
    if multi_session_vehicle_trend is not None and not multi_session_vehicle_trend.empty:
        _add_sheet(wb, "Vehicle Session Trend", multi_session_vehicle_trend, "Overall Assessment")
    if multi_session_transition_summary is not None and not multi_session_transition_summary.empty:
        _add_sheet(wb, "Session Transitions", multi_session_transition_summary, "Overall Assessment")
    if multi_session_transition_ecus is not None and not multi_session_transition_ecus.empty:
        _add_sheet(wb, "Transition ECU Diff", multi_session_transition_ecus, "Diff Status")
    if multi_session_transition_fields is not None and not multi_session_transition_fields.empty:
        _add_sheet(wb, "Transition Field Diff", multi_session_transition_fields, "Severity")
    if multi_session_ecu_timeline is not None and not multi_session_ecu_timeline.empty:
        _add_sheet(wb, "ECU Timeline", multi_session_ecu_timeline)
    if multi_session_change_history is not None and not multi_session_change_history.empty:
        _add_sheet(wb, "ECU Change History", multi_session_change_history)
    if configuration_diff_summary is not None and not configuration_diff_summary.empty:
        _add_sheet(wb, "Configuration Diff Summary", configuration_diff_summary, "Overall Assessment")
    if configuration_diff_ecus is not None and not configuration_diff_ecus.empty:
        _add_sheet(wb, "ECU Differences", configuration_diff_ecus, "Diff Status")
    if configuration_diff_fields is not None and not configuration_diff_fields.empty:
        _add_sheet(wb, "Field Differences", configuration_diff_fields, "Severity")
        for category, sheet_name in (
            ("SOFTWARE", "Software Differences"),
            ("HARDWARE", "Hardware Differences"),
            ("RELEASE", "Release Differences"),
            ("PRESENCE", "Added Removed ECUs"),
        ):
            scoped = configuration_diff_fields[
                configuration_diff_fields["Category"].astype(str) == category
            ]
            if not scoped.empty:
                _add_sheet(wb, sheet_name, scoped, "Severity")
    if release_recommendation_rules:
        recommendation_rule_rows = []
        for section, values in release_recommendation_rules.items():
            if isinstance(values, dict):
                for key, value in values.items():
                    recommendation_rule_rows.append({
                        "Section": section, "Rule": key, "Value": str(value)
                    })
            elif isinstance(values, list):
                for index, value in enumerate(values, start=1):
                    recommendation_rule_rows.append({
                        "Section": section, "Rule": index, "Value": str(value)
                    })
            else:
                recommendation_rule_rows.append({
                    "Section": "General", "Rule": section, "Value": str(values)
                })
        _add_sheet(wb, "Recommendation Rules", pd.DataFrame(recommendation_rule_rows))
    if release_path_rules:
        release_rule_rows = []
        for section, values in release_path_rules.items():
            if isinstance(values, dict):
                for key, value in values.items():
                    release_rule_rows.append({
                        "Section": section, "Rule": key, "Value": str(value)
                    })
            elif isinstance(values, list):
                for index, value in enumerate(values, start=1):
                    release_rule_rows.append({
                        "Section": section, "Rule": index, "Value": str(value)
                    })
            else:
                release_rule_rows.append({
                    "Section": "General", "Rule": section, "Value": str(values)
                })
        _add_sheet(wb, "Release Path Rules", pd.DataFrame(release_rule_rows))
    if update_impact_rules:
        update_rule_rows = []
        for section, values in update_impact_rules.items():
            if isinstance(values, dict):
                for key, value in values.items():
                    update_rule_rows.append({
                        "Section": section, "Rule": key, "Value": str(value)
                    })
            elif isinstance(values, list):
                for index, value in enumerate(values, start=1):
                    update_rule_rows.append({
                        "Section": section, "Rule": index, "Value": str(value)
                    })
            else:
                update_rule_rows.append({
                    "Section": "General", "Rule": section, "Value": str(values)
                })
        _add_sheet(wb, "Update Impact Rules", pd.DataFrame(update_rule_rows))
    if network_rules:
        network_rule_rows = []
        for section, values in network_rules.items():
            if isinstance(values, dict):
                for key, value in values.items():
                    network_rule_rows.append({
                        "Section": section, "Rule": key, "Value": str(value)
                    })
            elif isinstance(values, list):
                for index, value in enumerate(values, start=1):
                    network_rule_rows.append({
                        "Section": section, "Rule": index, "Value": str(value)
                    })
            else:
                network_rule_rows.append({
                    "Section": "General", "Rule": section, "Value": str(values)
                })
        _add_sheet(wb, "Network Rules", pd.DataFrame(network_rule_rows))
    _vehicle_report_sheets(wb, overview, details)

    if metadata.get("engineering_notes"):
        ws = wb.create_sheet("Engineering Notes")
        ws["A1"] = "Engineering Notes"
        ws["A1"].font = Font(size=16, bold=True, color=HEADER)
        ws["A3"] = metadata["engineering_notes"]
        ws["A3"].alignment = Alignment(wrap_text=True, vertical="top")
        ws.column_dimensions["A"].width = 110
        ws.row_dimensions[3].height = 240
        ws.sheet_view.showGridLines = False

    out = io.BytesIO()
    wb.save(out)
    data = out.getvalue()

    check = load_workbook(io.BytesIO(data), read_only=True, data_only=False)
    required = {
        "Executive Summary", "Dashboard Status", "Vehicle Summary", "Priority Actions",
        "Compliance Overview", "Release Details", "Field Details"
    }
    missing = required.difference(check.sheetnames)
    check.close()
    if missing:
        raise ValueError(f"Generated report is incomplete: {sorted(missing)}")
    return data
