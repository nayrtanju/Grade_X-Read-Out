from __future__ import annotations

from io import BytesIO
from typing import Any, Mapping, Sequence

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import (
    PageBreak,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)


SECTION_ORDER = [
    "EXECUTIVE_SUMMARY",
    "COMPLIANCE",
    "ECU_DETAILS",
    "DTC",
    "RISK_HEALTH",
    "RELEASE_CONSISTENCY",
    "NETWORK",
    "UPDATE_PLANNING",
    "CONFIGURATION_DIFF",
    "MULTI_SESSION",
    "ADVANCED_SEARCH",
    "QUALITY_DASHBOARD",
    "RELEASE_COVERAGE",
    "OEM_AUDIT",
    "SESSION_MERGE",
    "DATA_QUALITY",
    "FLEET_SNAPSHOT",
    "FULL_ASSESSMENT",
    "PROGRAMMING_VALIDATION",
    "CORRECTIVE_ACTIONS",
    "CLOSURE_VERIFICATION",
]


def available_sections(
    section_frames: Mapping[str, Sequence[tuple[str, pd.DataFrame]]],
) -> list[str]:
    result: list[str] = []
    for section in SECTION_ORDER:
        frames = section_frames.get(section, [])
        if any(frame is not None and not frame.empty for _, frame in frames):
            result.append(section)
    return result


def report_manifest(
    selected_sections: Sequence[str],
    section_frames: Mapping[str, Sequence[tuple[str, pd.DataFrame]]],
) -> pd.DataFrame:
    rows = []
    for order, section in enumerate(selected_sections, start=1):
        frames = section_frames.get(section, [])
        populated = [
            (name, len(frame), len(frame.columns))
            for name, frame in frames
            if frame is not None and not frame.empty
        ]
        rows.append({
            "Order": order,
            "Section ID": section,
            "Tables": len(populated),
            "Total Rows": sum(item[1] for item in populated),
            "Table Names": ", ".join(item[0] for item in populated),
        })
    return pd.DataFrame(rows)


def _safe_sheet_name(name: str, used: set[str]) -> str:
    cleaned = (
        str(name)
        .replace("/", "-")
        .replace("\\", "-")
        .replace("*", "")
        .replace("?", "")
        .replace("[", "(")
        .replace("]", ")")
        .replace(":", "-")
    )[:31] or "Sheet"
    candidate = cleaned
    index = 2
    while candidate in used:
        suffix = f" {index}"
        candidate = f"{cleaned[:31-len(suffix)]}{suffix}"
        index += 1
    used.add(candidate)
    return candidate


def _write_frame(ws, frame: pd.DataFrame) -> None:
    if frame is None or frame.empty:
        ws["A1"] = "No data available."
        return

    for column_index, column in enumerate(frame.columns, start=1):
        cell = ws.cell(row=1, column=column_index, value=str(column))
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_index, row in enumerate(frame.itertuples(index=False, name=None), start=2):
        for column_index, value in enumerate(row, start=1):
            if pd.isna(value):
                value = ""
            ws.cell(row=row_index, column=column_index, value=value)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for column_index, column in enumerate(frame.columns, start=1):
        values = [str(column)]
        values.extend(
            str(value)
            for value in frame.iloc[:, column_index - 1].head(200).fillna("")
        )
        width = min(60, max(10, max(len(value) for value in values) + 2))
        ws.column_dimensions[get_column_letter(column_index)].width = width


def build_dynamic_excel(
    *,
    metadata: Mapping[str, Any],
    selected_sections: Sequence[str],
    section_frames: Mapping[str, Sequence[tuple[str, pd.DataFrame]]],
) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)
    used: set[str] = set()

    cover = wb.create_sheet(_safe_sheet_name("Report Cover", used))
    cover["A1"] = metadata.get("title", "Grade-X Dynamic Engineering Report")
    cover["A1"].font = Font(size=18, bold=True)
    cover["A3"] = "Subtitle"
    cover["B3"] = metadata.get("subtitle", "")
    cover["A4"] = "Prepared By"
    cover["B4"] = metadata.get("prepared_by", "")
    cover["A5"] = "Project / Department"
    cover["B5"] = metadata.get("project", "")
    cover["A6"] = "Report Profile"
    cover["B6"] = metadata.get("profile", "")
    cover["A7"] = "Executive Comment"
    cover["B7"] = metadata.get("executive_comment", "")
    cover["A8"] = "Engineering Notes"
    cover["B8"] = metadata.get("engineering_notes", "")
    cover.column_dimensions["A"].width = 24
    cover.column_dimensions["B"].width = 90
    for row in range(3, 9):
        cover[f"A{row}"].font = Font(bold=True)
        cover[f"B{row}"].alignment = Alignment(wrap_text=True, vertical="top")

    manifest = report_manifest(selected_sections, section_frames)
    ws_manifest = wb.create_sheet(_safe_sheet_name("Report Manifest", used))
    _write_frame(ws_manifest, manifest)

    for section in selected_sections:
        for table_name, frame in section_frames.get(section, []):
            if frame is None or frame.empty:
                continue
            ws = wb.create_sheet(_safe_sheet_name(table_name, used))
            ws["A1"] = f"{section.replace('_', ' ').title()} — {table_name}"
            ws["A1"].font = Font(size=14, bold=True)
            for column_index, column in enumerate(frame.columns, start=1):
                cell = ws.cell(row=3, column=column_index, value=str(column))
                cell.font = Font(bold=True)
                cell.fill = PatternFill("solid", fgColor="D9EAF7")
            for row_index, row in enumerate(
                frame.itertuples(index=False, name=None), start=4
            ):
                for column_index, value in enumerate(row, start=1):
                    ws.cell(
                        row=row_index,
                        column=column_index,
                        value="" if pd.isna(value) else value,
                    )
            ws.freeze_panes = "A4"
            ws.auto_filter.ref = f"A3:{get_column_letter(len(frame.columns))}{len(frame)+3}"
            for column_index, column in enumerate(frame.columns, start=1):
                sample = [str(column)] + [
                    str(value)
                    for value in frame.iloc[:, column_index - 1]
                    .head(200).fillna("")
                ]
                ws.column_dimensions[get_column_letter(column_index)].width = min(
                    60, max(10, max(len(value) for value in sample) + 2)
                )

    output = BytesIO()
    wb.save(output)
    return output.getvalue()


def _pdf_table(frame: pd.DataFrame, styles, max_rows: int = 60) -> Table:
    display = frame.head(max_rows).copy()
    if display.empty:
        display = pd.DataFrame([{"Result": "No data available"}])
    display = display.iloc[:, :12]
    data = [[Paragraph(str(column), styles["ReportTableHeader"]) for column in display.columns]]
    for row in display.itertuples(index=False, name=None):
        data.append([
            Paragraph(str("" if pd.isna(value) else value), styles["ReportTableCell"])
            for value in row
        ])
    page_width = landscape(A4)[0] - 24 * mm
    widths = [page_width / len(display.columns)] * len(display.columns)
    table = Table(data, colWidths=widths, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#D9EAF7")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 3),
        ("RIGHTPADDING", (0, 0), (-1, -1), 3),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]))
    return table


def build_dynamic_pdf(
    *,
    metadata: Mapping[str, Any],
    selected_sections: Sequence[str],
    section_frames: Mapping[str, Sequence[tuple[str, pd.DataFrame]]],
) -> bytes:
    output = BytesIO()
    doc = SimpleDocTemplate(
        output,
        pagesize=landscape(A4),
        rightMargin=12 * mm,
        leftMargin=12 * mm,
        topMargin=12 * mm,
        bottomMargin=12 * mm,
        title=str(metadata.get("title", "Grade-X Dynamic Engineering Report")),
    )
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(
        name="ReportTitle",
        parent=styles["Title"],
        fontSize=22,
        leading=26,
        alignment=TA_CENTER,
        spaceAfter=12,
    ))
    styles.add(ParagraphStyle(
        name="ReportSection",
        parent=styles["Heading1"],
        fontSize=16,
        leading=19,
        spaceAfter=8,
    ))
    styles.add(ParagraphStyle(
        name="ReportTableHeader",
        parent=styles["Normal"],
        fontSize=6.5,
        leading=8,
        fontName="Helvetica-Bold",
    ))
    styles.add(ParagraphStyle(
        name="ReportTableCell",
        parent=styles["Normal"],
        fontSize=6,
        leading=7,
    ))

    story = [
        Spacer(1, 25 * mm),
        Paragraph(str(metadata.get("title", "Grade-X Dynamic Engineering Report")), styles["ReportTitle"]),
        Paragraph(str(metadata.get("subtitle", "")), styles["Heading2"]),
        Spacer(1, 10 * mm),
        Paragraph(f"<b>Prepared by:</b> {metadata.get('prepared_by', '')}", styles["Normal"]),
        Paragraph(f"<b>Project / Department:</b> {metadata.get('project', '')}", styles["Normal"]),
        Paragraph(f"<b>Report profile:</b> {metadata.get('profile', '')}", styles["Normal"]),
        Spacer(1, 6 * mm),
        Paragraph(f"<b>Executive comment:</b> {metadata.get('executive_comment', '')}", styles["Normal"]),
        Spacer(1, 3 * mm),
        Paragraph(f"<b>Engineering notes:</b> {metadata.get('engineering_notes', '')}", styles["Normal"]),
        PageBreak(),
        Paragraph("Report Manifest", styles["ReportSection"]),
        _pdf_table(report_manifest(selected_sections, section_frames), styles),
    ]

    for section in selected_sections:
        populated = [
            (name, frame)
            for name, frame in section_frames.get(section, [])
            if frame is not None and not frame.empty
        ]
        if not populated:
            continue
        story.append(PageBreak())
        story.append(Paragraph(
            section.replace("_", " ").title(),
            styles["ReportSection"],
        ))
        for table_name, frame in populated:
            story.append(Paragraph(str(table_name), styles["Heading2"]))
            story.append(_pdf_table(frame, styles))
            story.append(Spacer(1, 6 * mm))

    doc.build(story)
    return output.getvalue()
