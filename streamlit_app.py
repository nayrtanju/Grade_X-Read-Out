rom __future__ import annotations

import io
import re
import xml.etree.ElementTree as ET
from dataclasses import asdict, dataclass
from typing import Iterable, Optional

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


st.set_page_config(
    page_title="Grade-X Session Analyzer",
    page_icon="🔧",
    layout="wide",
)


def clean(value: Optional[str]) -> str:
    return (value or "").strip()


def norm(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", value.lower())


def natural_join(prefix: str, number: str, revision: str) -> str:
    prefix, number, revision = clean(prefix), clean(number), clean(revision)
    if not any((prefix, number, revision)):
        return ""
    base = f"{prefix}{number}".strip()
    return f"{base} Rev.{revision}" if revision else base


@dataclass
class EcuRow:
    source_file: str
    session_id: str
    vehicle_model: str
    manufacturer: str
    ecu_id: str
    ecu_name: str
    vin: str
    hardware_number: str
    application_sw: str
    calibration_sw: str
    basic_sw: str
    bootloader: str
    identification_count: int
    dtc_count: int


def identification_records(ecu: ET.Element) -> list[dict[str, str]]:
    records: list[dict[str, str]] = []
    for item in ecu.findall("identifications"):
        records.append(
            {
                "id": clean(item.findtext("id")),
                "display": clean(item.findtext("displayText")),
                "value": clean(item.findtext("value")),
            }
        )
    return records


def find_value(
    records: list[dict[str, str]],
    *,
    include_all: Iterable[str],
    exclude_any: Iterable[str] = (),
) -> str:
    includes = [norm(x) for x in include_all]
    excludes = [norm(x) for x in exclude_any]

    def matches(text: str) -> bool:
        value = norm(text)
        return all(token in value for token in includes) and not any(
            token in value for token in excludes
        )

    for field in ("id", "display"):
        for record in records:
            if record["value"] and matches(record[field]):
                return record["value"]
    return ""


def category_match(text: str, category: str) -> bool:
    value = norm(text)
    if category == "bootloader":
        return "bootloader" in value or ("boot" in value and "loader" in value)
    return category in value and ("software" in value or "sw" in value)


def software_value(records: list[dict[str, str]], category: str) -> str:
    candidates = [
        record
        for record in records
        if category_match(record["id"] + " " + record["display"], category)
    ]

    prefix = ""
    number = ""
    revision = ""

    for record in candidates:
        text = norm(record["id"] + " " + record["display"])
        value = record["value"]
        if not value or "frs" in text:
            continue
        if "prefix" in text and not prefix:
            prefix = value
        elif "revision" in text and not revision:
            revision = value
        elif "number" in text and "prefix" not in text and "revision" not in text:
            if not number or re.search(r"\d", value):
                number = value

    numeric_numbers = []
    for record in candidates:
        text = norm(record["id"] + " " + record["display"])
        value = record["value"]
        if (
            value
            and "number" in text
            and "prefix" not in text
            and "revision" not in text
            and "frs" not in text
            and re.search(r"\d", value)
        ):
            numeric_numbers.append(value)
    if numeric_numbers:
        number = numeric_numbers[0]

    return natural_join(prefix, number, revision)


def extract_vin(records: list[dict[str, str]], session_id: str) -> str:
    vin = find_value(records, include_all=["vehicle", "identification", "number"])
    if vin:
        return vin
    candidate = session_id.split("@", 1)[0].strip()
    return candidate if re.fullmatch(r"[A-HJ-NPR-Z0-9]{17}", candidate, re.I) else ""


def extract_hardware(records: list[dict[str, str]]) -> str:
    candidates = [
        record
        for record in records
        if "hardware" in norm(record["id"] + " " + record["display"])
        and "software" not in norm(record["id"] + " " + record["display"])
    ]

    prefix = ""
    number = ""
    revision = ""
    for record in candidates:
        text = norm(record["id"] + " " + record["display"])
        value = record["value"]
        if not value:
            continue
        if "prefix" in text and not prefix:
            prefix = value
        elif "revision" in text and not revision:
            revision = value
        elif "number" in text and "prefix" not in text and "revision" not in text:
            if "supplier" not in text or not number:
                number = value

    if number:
        return natural_join(prefix, number, revision)

    part_candidates = [
        record
        for record in records
        if "ecupartnumber" in norm(record["id"] + " " + record["display"])
        and not any(
            token in norm(record["id"] + " " + record["display"])
            for token in ("application", "calibration", "basic", "software")
        )
    ]
    for record in part_candidates:
        text = norm(record["id"] + " " + record["display"])
        value = record["value"]
        if not value:
            continue
        if "prefix" in text and not prefix:
            prefix = value
        elif "revision" in text and not revision:
            revision = value
        elif "prefix" not in text and "revision" not in text and re.search(r"\d", value):
            number = value

    return natural_join(prefix, number, revision)


def parse_session_bytes(file_name: str, data: bytes) -> tuple[list[EcuRow], list[dict[str, str]]]:
    try:
        root = ET.fromstring(data)
    except ET.ParseError as exc:
        raise ValueError(f"Geçersiz XML/session yapısı: {exc}") from exc

    session_id = clean(root.findtext("id"))
    vehicle_model = clean(root.findtext("./machine/model"))
    manufacturer = clean(root.findtext("./machine/manufacturer"))

    rows: list[EcuRow] = []
    raw_rows: list[dict[str, str]] = []

    ecus = root.findall("./machine/networks/ecus")
    if not ecus:
        raise ValueError("Dosyada './machine/networks/ecus' düğümleri bulunamadı.")

    for ecu in ecus:
        ecu_id = clean(ecu.findtext("id"))
        ecu_name = clean(ecu.findtext("displayName"))
        records = identification_records(ecu)

        for record in records:
            raw_rows.append(
                {
                    "Source File": file_name,
                    "Session ID": session_id,
                    "ECU ID": ecu_id,
                    "ECU Name": ecu_name,
                    "Display Text": record["display"],
                    "Value": record["value"],
                    "Identifier ID": record["id"],
                }
            )

        rows.append(
            EcuRow(
                source_file=file_name,
                session_id=session_id,
                vehicle_model=vehicle_model,
                manufacturer=manufacturer,
                ecu_id=ecu_id,
                ecu_name=ecu_name,
                vin=extract_vin(records, session_id),
                hardware_number=extract_hardware(records),
                application_sw=software_value(records, "application"),
                calibration_sw=software_value(records, "calibration"),
                basic_sw=software_value(records, "basic"),
                bootloader=software_value(records, "bootloader"),
                identification_count=len(records),
                dtc_count=len(ecu.findall("dtcs")),
            )
        )

    return rows, raw_rows


def rows_to_dataframe(rows: list[EcuRow]) -> pd.DataFrame:
    column_map = {
        "source_file": "Source File",
        "session_id": "Session ID",
        "vehicle_model": "Vehicle Model",
        "manufacturer": "Manufacturer",
        "ecu_id": "ECU ID",
        "ecu_name": "ECU Name",
        "vin": "VIN",
        "hardware_number": "Hardware Number",
        "application_sw": "Application SW",
        "calibration_sw": "Calibration SW",
        "basic_sw": "Basic SW",
        "bootloader": "Bootloader",
        "identification_count": "Identifier Count",
        "dtc_count": "DTC Count",
    }
    frame = pd.DataFrame([asdict(row) for row in rows])
    return frame.rename(columns=column_map)


def autosize_sheet(worksheet, max_width: int = 55) -> None:
    for column_cells in worksheet.columns:
        width = 0
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            width = max(width, len(value))
        worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(
            width + 2, max_width
        )


def dataframe_to_sheet(workbook: Workbook, title: str, dataframe: pd.DataFrame) -> None:
    worksheet = workbook.create_sheet(title)
    worksheet.append(list(dataframe.columns))
    for row in dataframe.itertuples(index=False, name=None):
        worksheet.append(list(row))

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    if worksheet.max_row >= 2 and worksheet.max_column >= 1:
        table_ref = f"A1:{get_column_letter(worksheet.max_column)}{worksheet.max_row}"
        safe_name = re.sub(r"[^A-Za-z0-9_]", "", title.replace(" ", "")) + "Table"
        table = Table(displayName=safe_name, ref=table_ref)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        worksheet.add_table(table)

    for row in worksheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    autosize_sheet(worksheet)


def create_excel(summary_df: pd.DataFrame, raw_df: pd.DataFrame) -> bytes:
    workbook = Workbook()
    workbook.remove(workbook.active)
    dataframe_to_sheet(workbook, "ECU Summary", summary_df)
    dataframe_to_sheet(workbook, "Raw Identifiers", raw_df)

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return output.getvalue()


st.title("Bosch Grade-X Session Analyzer")
st.caption(
    "Grade-X `.session` dosyalarından ECU yazılım, kalibrasyon, bootloader, VIN ve hardware bilgilerini çıkarır."
)

uploaded_files = st.file_uploader(
    "Bir veya birden fazla .session dosyası yükleyin",
    type=["session", "xml"],
    accept_multiple_files=True,
)

if not uploaded_files:
    st.info("Analizi başlatmak için en az bir `.session` dosyası yükleyin.")
    st.stop()

all_rows: list[EcuRow] = []
all_raw_rows: list[dict[str, str]] = []
errors: list[str] = []

for uploaded_file in uploaded_files:
    try:
        rows, raw_rows = parse_session_bytes(uploaded_file.name, uploaded_file.getvalue())
        all_rows.extend(rows)
        all_raw_rows.extend(raw_rows)
    except Exception as exc:
        errors.append(f"{uploaded_file.name}: {exc}")

if errors:
    for error in errors:
        st.error(error)

if not all_rows:
    st.warning("Geçerli bir ECU kaydı çıkarılamadı.")
    st.stop()

summary_df = rows_to_dataframe(all_rows)
raw_df = pd.DataFrame(all_raw_rows)

with st.sidebar:
    st.header("Filtreler")
    available_ecus = sorted(summary_df["ECU ID"].dropna().astype(str).unique())
    selected_ecus = st.multiselect("ECU", available_ecus, default=available_ecus)
    only_with_identifiers = st.checkbox("Sadece identifier bulunan ECU'lar", value=False)
    only_with_dtc = st.checkbox("Sadece DTC bulunan ECU'lar", value=False)

filtered_df = summary_df[summary_df["ECU ID"].isin(selected_ecus)].copy()
if only_with_identifiers:
    filtered_df = filtered_df[filtered_df["Identifier Count"] > 0]
if only_with_dtc:
    filtered_df = filtered_df[filtered_df["DTC Count"] > 0]

metric_1, metric_2, metric_3, metric_4 = st.columns(4)
metric_1.metric("Session dosyası", len(uploaded_files))
metric_2.metric("Toplam ECU", len(summary_df))
metric_3.metric("Identifier bulunan ECU", int((summary_df["Identifier Count"] > 0).sum()))
metric_4.metric("DTC bulunan ECU", int((summary_df["DTC Count"] > 0).sum()))

summary_tab, raw_tab = st.tabs(["ECU Summary", "Raw Identifiers"])

with summary_tab:
    st.dataframe(
        filtered_df,
        use_container_width=True,
        hide_index=True,
        column_config={
            "Identifier Count": st.column_config.NumberColumn(format="%d"),
            "DTC Count": st.column_config.NumberColumn(format="%d"),
        },
    )

with raw_tab:
    selected_raw_ecus = set(filtered_df["ECU ID"].astype(str))
    filtered_raw_df = raw_df[raw_df["ECU ID"].astype(str).isin(selected_raw_ecus)]
    st.dataframe(filtered_raw_df, use_container_width=True, hide_index=True)

excel_bytes = create_excel(summary_df, raw_df)
st.download_button(
    label="Excel raporunu indir",
    data=excel_bytes,
    file_name="GradeX_ECU_Identifiers.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    type="primary",
)
